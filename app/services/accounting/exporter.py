"""
Exporter: converts BillRecord objects into CSV rows and XLSX workbooks.

Turkish column names are used in all exported files per product requirements.
"""

from __future__ import annotations

import csv
import io
import json
import re
from pathlib import Path
from typing import Mapping, Optional, Sequence

from app.models.schemas import BillRecord
from app.utils.logging import get_logger

logger = get_logger(__name__)

_HYPERLINK_FORMULA_RE = re.compile(r'^=HYPERLINK\("(?P<url>[^"]+)"[;,]"(?P<label>[^"]*)"\)$')


def _hyperlink_parts(value: object) -> tuple[str, str] | None:
    raw = str(value or '').strip()
    if not raw:
        return None
    match = _HYPERLINK_FORMULA_RE.match(raw)
    if match:
        return match.group('url'), match.group('label') or match.group('url')
    if raw.startswith('http://') or raw.startswith('https://'):
        return raw, raw
    return None

# Ordered mapping: internal field → Turkish export column
COLUMN_MAP: dict[str, str] = {
    "company_name": "Firma Adı",
    "tax_number": "Vergi Numarası",
    "tax_office": "Vergi Dairesi",
    "document_number": "Belge Numarası",
    "invoice_number": "Fatura Numarası",
    "receipt_number": "Fiş Numarası",
    "document_date": "Tarih",
    "document_time": "Saat",
    "currency": "Para Birimi",
    "subtotal": "Ara Toplam",
    "vat_rate": "KDV Oranı",
    "vat_amount": "KDV Tutarı",
    "total_amount": "Genel Toplam",
    "sender_name": "Gönderen Adı",
    "recipient_name": "Alıcı / Tedarikçi",
    "buyer_name": "Alıcı",
    "invoice_type": "Fatura Tipi",
    "line_quantity": "Miktar",
    "line_unit": "Birim",
    "unit_price": "Birim Fiyat",
    "line_amount": "Kalem Tutarı",
    "withholding_present": "Tevkifat Var mı",
    "withholding_rate": "Tevkifat Oranı",
    "withholding_amount": "Tevkifat Tutarı",
    "payable_amount": "Ödenecek Tutar",
    "iban": "IBAN",
    "bank_name": "Banka",
    "shipment_origin": "Çıkış Yeri",
    "shipment_destination": "Sevk Yeri",
    "pallet_count": "Palet Sayısı",
    "items_per_pallet": "Adet/Palet",
    "product_quantity": "Ürün Miktarı",
    "vehicle_plate": "Plaka",
    "cheque_issue_place": "Çek Keşide Yeri",
    "cheque_issue_date": "Çek Keşide Tarihi",
    "cheque_due_date": "Çek Vade Tarihi",
    "cheque_serial_number": "Çek Seri No",
    "cheque_bank_name": "Çek Banka",
    "cheque_branch": "Çek Şube",
    "cheque_account_ref": "Çek Hesap Ref",
    "line_items": "Fatura Kalemleri JSON",
    "payment_method": "Ödeme Yöntemi",
    "expense_category": "Gider Kategorisi",
    "description": "Açıklama",
    "notes": "Notlar",
    "source_message_id": "Kaynak Mesaj ID",
    "source_filename": "Kaynak Dosya Adı",
    "source_type": "Kaynak Türü",
    "source_sender_id": "Kaynak Gönderen ID",
    "source_sender_name": "Kaynak Gönderen Adı",
    "source_group_id": "Kaynak Grup ID",
    "source_chat_type": "Sohbet Türü",
    "confidence": "Güven Skoru",
}

TURKISH_HEADERS = list(COLUMN_MAP.values())


def record_to_row(record: BillRecord) -> dict[str, str]:
    """Convert a BillRecord to a flat dict with Turkish column names."""
    raw = record.model_dump()
    row: dict[str, str] = {}
    for field, column in COLUMN_MAP.items():
        value = raw.get(field)
        if value is None:
            row[column] = ""
        elif field == "line_items":
            row[column] = json.dumps(value, ensure_ascii=False)
        else:
            row[column] = str(value)
    return row


def records_to_csv(records: Sequence[BillRecord]) -> str:
    """Return a UTF-8 CSV string (with BOM for Excel compatibility)."""
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=TURKISH_HEADERS, extrasaction="ignore")
    writer.writeheader()
    for record in records:
        writer.writerow(record_to_row(record))
    return "\ufeff" + output.getvalue()  # BOM for Excel


def save_csv(records: Sequence[BillRecord], filepath: Path) -> Path:
    """Write records to a CSV file and return the Path."""
    filepath.parent.mkdir(parents=True, exist_ok=True)
    filepath.write_text(records_to_csv(records), encoding="utf-8-sig")
    logger.info("CSV saved: %s (%d records)", filepath, len(records))
    return filepath


def records_to_xlsx_bytes(records: Sequence[BillRecord]) -> bytes:
    """
    Return an XLSX workbook as bytes.

    Requires openpyxl (already in requirements.txt).
    """
    rows = [record_to_row(record) for record in records]
    return tabular_rows_to_xlsx_bytes(rows, TURKISH_HEADERS)


def tabular_rows_to_xlsx_bytes(
    rows: Sequence[Mapping[str, str]], headers: Sequence[str] | None = None
) -> bytes:
    """
    Return an XLSX workbook from already-tabular row data.

    `rows` should be keyed by header name.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError as exc:
        raise ImportError("openpyxl is required for XLSX export.") from exc

    workbook_headers = list(headers or TURKISH_HEADERS)
    wb = Workbook()
    ws = wb.active
    ws.title = "Muhasebe"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E79")

    for col_idx, header in enumerate(workbook_headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, header in enumerate(workbook_headers, start=1):
            value = row.get(header, "")
            cell = ws.cell(row=row_idx, column=col_idx)
            hyperlink = _hyperlink_parts(value) if header == "Belge" else None
            if hyperlink:
                url, label = hyperlink
                cell.value = label
                cell.hyperlink = url
                cell.style = "Hyperlink"
            else:
                cell.value = value

    # Auto-fit column widths (approximate)
    for col in ws.columns:
        max_length = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def save_xlsx(records: Sequence[BillRecord], filepath: Path) -> Path:
    """Write records to an XLSX file and return the Path."""
    filepath.parent.mkdir(parents=True, exist_ok=True)
    filepath.write_bytes(records_to_xlsx_bytes(records))
    logger.info("XLSX saved: %s (%d records)", filepath, len(records))
    return filepath
