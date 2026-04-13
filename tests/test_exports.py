"""
Tests for export download endpoints.
"""

from __future__ import annotations

from io import BytesIO
from pathlib import Path
from tempfile import TemporaryDirectory
from unittest.mock import patch

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.main import app
from app.services.accounting.exporter import tabular_rows_to_xlsx_bytes


def test_export_csv_returns_latest_export_file():
    with TemporaryDirectory() as tmpdir:
        export_dir = Path(tmpdir) / "exports"
        export_dir.mkdir(parents=True, exist_ok=True)
        filepath = export_dir / "records_2026-04-03.csv"
        filepath.write_text("Firma Adı,Genel Toplam\nABC Market,100.0\n", encoding="utf-8-sig")

        client = TestClient(app)
        with patch("app.main.settings.storage_dir", tmpdir):
            response = client.get("/export.csv")

    assert response.status_code == 200
    assert "text/csv" in response.headers["content-type"]
    assert "ABC Market" in response.text


def test_export_xlsx_returns_downloadable_workbook():
    with TemporaryDirectory() as tmpdir:
        export_dir = Path(tmpdir) / "exports"
        export_dir.mkdir(parents=True, exist_ok=True)
        filepath = export_dir / "records_2026-04-03.csv"
        filepath.write_text("Firma Adı,Genel Toplam\nABC Market,100.0\n", encoding="utf-8-sig")

        client = TestClient(app)
        with patch("app.main.settings.storage_dir", tmpdir):
            response = client.get("/export.xlsx")

    assert response.status_code == 200
    assert (
        response.headers["content-type"]
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert response.headers["content-disposition"].endswith('"records_2026-04-03.xlsx"')
    assert response.content[:2] == b"PK"


def test_export_endpoints_return_404_when_no_export_exists():
    with TemporaryDirectory() as tmpdir:
        client = TestClient(app)
        with patch("app.main.settings.storage_dir", tmpdir):
            csv_response = client.get("/export.csv")
            xlsx_response = client.get("/export.xlsx")

    assert csv_response.status_code == 404
    assert xlsx_response.status_code == 404


def test_tabular_xlsx_export_preserves_belge_as_clickable_hyperlink():
    workbook_bytes = tabular_rows_to_xlsx_bytes(
        [{
            "Firma Adı": "ABC Market",
            "Belge": '=HYPERLINK("https://drive.google.com/file/d/test/view";"Görüntüle")',
        }],
        headers=["Firma Adı", "Belge"],
    )

    workbook = load_workbook(BytesIO(workbook_bytes))
    sheet = workbook.active
    cell = sheet.cell(row=2, column=2)

    assert cell.value == "Görüntüle"
    assert cell.hyperlink is not None
    assert cell.hyperlink.target == "https://drive.google.com/file/d/test/view"

